#!/usr/bin/env python3
"""Обзвон по свежим коммерческим пермитам Fort Lauderdale (Broward):
GC + телефон подрядчика + объект + тип работ → прямой звонок (final clean / отделка-суб)."""
import json, datetime, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SP = "/tmp/claude-0/-home-user-Supply-of-Goods/4f5e4e6d-222f-57da-93ad-94f5b2132f9f/scratchpad"
def load(f):
    d = json.load(open(os.path.join(SP, f)))
    return [x["attributes"] for x in d.get("features", [])]

feats = load("ftl3.json") + load("ftl4.json")

def dt(v):
    return datetime.datetime.utcfromtimestamp(v/1000).date().isoformat() if v else ""

def clean_ph(p):
    p = (p or "").strip()
    digs = "".join(c for c in p if c.isdigit())
    if len(digs) < 10 or digs == "0000000000":
        return ""
    return f"({digs[-10:-7]}) {digs[-7:-4]}-{digs[-4:]}"

def offer(desc, uc):
    d = (desc or "").lower(); u = (uc or "").lower()
    if any(k in u for k in ("restaurant","bakery","bar","coffee","cafe")) or "outdoor dining" in d or "change of use" in d:
        return "Отделка-суб + FINAL CLEAN (свежий buildout — большая уборка)"
    if "new" in d:
        return "FINAL CLEAN (ground-up сдача) + отделка-суб"
    return "FINAL CLEAN после ремонта + отделка-суб (paint/floor/stucco)"

rows = []
seen = set()
for a in feats:
    ph = clean_ph(a.get("CONTRACTPH"))
    gc = (a.get("CONTRACTOR") or "").strip()
    if not ph or not gc:
        continue
    key = (gc, ph)
    if key in seen:
        continue
    seen.add(key)
    rows.append({
        "gc": gc.title(),
        "phone": ph,
        "addr": (a.get("FULLADDR") or "").strip(),
        "work": (a.get("PERMITDESC") or "").replace(" Permit", ""),
        "use": (a.get("USECLASS") or "").title() if a.get("USECLASS") not in (None,"None") else "",
        "status": a.get("PERMITSTAT") or "",
        "date": dt(a.get("SUBMITDT")),
        "owner": (a.get("OWNERNAME") or "").title() if a.get("OWNERNAME") else "",
        "offer": offer(a.get("PERMITDESC"), a.get("USECLASS")),
    })
rows.sort(key=lambda r: r["date"], reverse=True)

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Broward — обзвон GC"
BRAND = "2E4A62"
F = ["Подрядчик (GC)", "Телефон", "Объект (адрес)", "Тип работ", "Назначение",
     "Статус", "Дата подачи", "Владелец", "Что предложить в звонке", "Город"]
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor=BRAND)
thin = Side(style="thin", color="D0D0D0"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
ws.append(F)
for c in ws[1]:
    c.font = hf; c.fill = hfill; c.border = bd
    c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
food = PatternFill("solid", fgColor="FDE9D8")   # restaurant/change-of-use
new = PatternFill("solid", fgColor="E7F0E9")     # new build
for r in rows:
    ws.append([r["gc"], r["phone"], r["addr"], r["work"], r["use"], r["status"],
               r["date"], r["owner"], r["offer"], "Fort Lauderdale, FL"])
    row = ws[ws.max_row]
    fill = food if ("Отделка-суб + FINAL" in r["offer"]) else (new if r["offer"].startswith("FINAL CLEAN (ground") else None)
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = bd
        if fill: c.fill = fill
for i, wd in enumerate([28, 16, 24, 22, 18, 20, 12, 22, 40, 18], 1):
    ws.column_dimensions[get_column_letter(i)].width = wd
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(F))}{ws.max_row}"
out = os.path.join(os.path.dirname(__file__), "..", "reports", "broward-permit-calllist-2026-08-26.xlsx")
wb.save(out)
print("wrote", out, "·", len(rows), "GC с телефоном")
