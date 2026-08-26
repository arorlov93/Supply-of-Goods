#!/usr/bin/env python3
"""Переранжировка свежих пермитов Miami-Dade по НАШЕЙ услуге:
P1 = покраска+гипсокартон+полы (интерьерный ремонт/buildout)
P2 = уборка (new/shell/addition — final clean на сдаче)
P3 = демонтаж (demo-скоуп)
Активные (CO нет) + ≤180 дней + телефон. + JSON топ-P1 для веб-сверки."""
import json, os, datetime, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SP = "/tmp/claude-0/-home-user-Supply-of-Goods/4f5e4e6d-222f-57da-93ad-94f5b2132f9f/scratchpad"
TODAY = datetime.date(2026, 8, 26)
TARGET = {
    "RETAIL SALES": "Ритейл", "OFFICE - PROFESSIONAL BUILDINGS": "Офис", "OFFICE USE ONLY": "Офис",
    "OFFICE - SALES": "Офис", "RESTAURANT-CAFETERIA": "Ресторан",
    "RESTAURANT/CAFET/BAR/LOUNGE/NIGHT CLUB": "Ресторан/бар", "BAR/COCKTAIL LOUNGE/RESTAURANTS": "Ресторан/бар",
    "BAR-LOUNGE-NIGHT CLUB": "Бар", "BAKERY PLANT": "Пекарня", "BANQUET HALL": "Банкетный зал",
    "CONVENIENCE STORE": "Магазин у дома", "PACKAGE STORE": "Ликёр-магазин",
    "CLINIC/SANITARIUMS/HEALTH CENTERS": "Клиника", "BEAUTY SALON-BARBER SHOP": "Салон/барбершоп",
    "GYM/EXERCISE CLUB": "Спортзал", "COIN LAUNDRY-DRY CLEANING": "Прачечная/химчистка",
    "DAYCARE - KINDERGARTEN": "Детсад",
}
def d2(s):
    try: return datetime.date.fromisoformat(s[:10])
    except: return None
def phone_fmt(p):
    d = "".join(c for c in (p or "") if c.isdigit())
    if len(d) < 10 or d[-10:] == "0000000000": return ""
    return f"({d[-10:-7]}) {d[-7:-4]}-{d[-4:]}"

def classify(app, cmt):
    a = (app or "").upper(); c = (cmt or "").upper()
    demo = bool(re.search(r"\bDEMO\b|DEMOLISH|DEMOLITION", c)) and "NEW" not in c
    interior = ("ALTER - INTERIOR" in a) or any(k in c for k in
        ("REMODEL","BUILDOUT","BUILD OUT","RENO","TENANT","ALTERATION","ALT INT","INT ALT","CHANGE OF USE","CHANGE USE","LEGALIZATION","ESTABLISH USE","FINISH"))
    newb = any(k in a for k in ("NEW","SHELL","ADDITION"))
    if demo and not interior and not newb:
        return 3, "ДЕМОНТАЖ", "Демонтаж/вывоз (наш субподряд)"
    if interior:
        return 1, "ОТДЕЛКА", "Покраска + гипсокартон + полы (интерьерный ремонт)"
    if newb:
        return 2, "УБОРКА", "Final clean на сдаче (+ отделка позже)"
    return 2, "УБОРКА", "Final clean после работ"

feats = []
for off in (0, 1000):
    feats += [f["attributes"] for f in json.load(open(f"{SP}/md_{off}.json")).get("features", [])]

rows = []; seen = set()
for a in feats:
    use = a.get("ProposedUseDescription")
    if use not in TARGET: continue
    ph = phone_fmt(a.get("ContractorPhone"))
    gc = (a.get("ContractorName") or "").strip()
    dt = d2(a.get("PermitIssuedDate"))
    if not ph or not gc or not dt: continue
    key = (gc, ph, (a.get("PropertyAddress") or "").strip())
    if key in seen: continue
    seen.add(key)
    age = (TODAY - dt).days
    tier = "🟢 ≤30д" if age <= 30 else ("🟢 31-90д" if age <= 90 else "🟡 91-180д")
    pri, ptag, svc = classify(a.get("ApplicationTypeDescription"), a.get("DetailDescriptionComments"))
    rows.append({
        "pri": pri, "ptag": ptag, "svc": svc,
        "owner": (a.get("OwnerName") or "").title(), "gc": gc.title(), "phone": ph,
        "addr": (a.get("PropertyAddress") or "").title(), "use": TARGET[use],
        "work": (a.get("ApplicationTypeDescription") or "").title().replace("Alter", "Ремонт"),
        "cmt": (a.get("DetailDescriptionComments") or "").strip().title(),
        "date": dt.isoformat(), "age": age, "tier": tier,
    })
# sort: priority asc, then freshest
rows.sort(key=lambda r: (r["pri"], r["age"]))

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Обзвон по приоритету"
BRAND = "2E4A62"
F = ["Приоритет", "Наша услуга", "Заказчик", "Подрядчик (GC)", "Телефон GC", "Объект (адрес)",
     "Назначение", "Тип работ", "Описание", "Дата", "Актуальность", "Веб-сверка №"]
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor=BRAND)
thin = Side(style="thin", color="D0D0D0"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
ws.append(F)
for c in ws[1]:
    c.font = hf; c.fill = hfill; c.border = bd
    c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
col = {1: PatternFill("solid", fgColor="DCE9F7"), 2: PatternFill("solid", fgColor="FDE9D8"), 3: PatternFill("solid", fgColor="ECECEC")}
for r in rows:
    ws.append([f"P{r['pri']} {r['ptag']}", r["svc"], r["owner"], r["gc"], r["phone"], r["addr"],
               r["use"], r["work"], r["cmt"], r["date"], r["tier"], ""])
    row = ws[ws.max_row]
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = bd; c.fill = col[r["pri"]]
for i, wd in enumerate([13, 34, 22, 25, 15, 24, 15, 16, 20, 12, 13, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = wd
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(F))}{ws.max_row}"
out = os.path.join(os.path.dirname(__file__), "..", "reports", "permit-priority-2026-08-26.xlsx")
wb.save(out)

from collections import Counter
pc = Counter(r["pri"] for r in rows)
print(f"wrote {out} · всего {len(rows)} | P1 отделка:{pc[1]} P2 уборка:{pc[2]} P3 демонтаж:{pc[3]}")
# top P1 freshest for web-verify (<=90 days)
p1 = [r for r in rows if r["pri"] == 1 and r["age"] <= 90]
p1.sort(key=lambda r: r["age"])
verify = [{"gc": r["gc"], "phone": r["phone"], "addr": r["addr"], "date": r["date"]} for r in p1[:45]]
json.dump(verify, open(f"{SP}/verify_p1.json", "w"))
print(f"P1 ≤90д для веб-сверки: {len(p1)} (беру топ {len(verify)})")
