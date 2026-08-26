#!/usr/bin/env python3
"""Excel: провайдеры кредитных ЛИНИЙ под недвижимость (revolving), с фокусом «новичок vs опыт»."""
import openpyxl, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F = ["Провайдер", "Тип линии", "Размер линии", "Возобновляемая?", "Кому дают (квалификация)",
     "Новичку без истории?", "LLC + без дохода", "Контакт", "Сайт"]
rows = [
    # доступно новичку сейчас
    ("FundingPilot", "Unsecured business LOC (stated income)", "$10k – $250k", "Да (revolving счёт)", "FICO 680+ (под 700 — со-подписант), без collateral", "✅ ДА — прямо для первых флипов", "Да", "(888) 860-2844", "fundingpilot.com/fix-flip-credit-lines-stated-income", 1),
    ("Investor HELOC (Hurst 'Equity Leverage Line' / банки)", "HELOC под equity инвест-объекта", "по equity (до ~80–100% LTC)", "Да (линия, платёж только при draw)", "FICO 680–700+, НУЖЕН существующий объект с equity", "🟡 только если есть недвижимость с equity", "LLC ок (зависит от банка)", "Hurst (877) 292-7350", "hurstlending.com/investor-loans/investment-property-heloc", 1),
    # настоящие flip-линии, но нужен опыт
    ("LendingOne", "Fix&flip LINE (12-мес runway)", "до ~$3M", "Да (recycle между сделками)", "Для ОПЫТНЫХ (несколько проектов); новичка — на per-deal", "❌ нет (новичку per-deal)", "Да", "(888) 987-1276 · Boca Raton FL", "lendingone.com/loan-types/fix-and-flip-loans", 2),
    ("Lima One Capital", "Fix&flip / new-constr LINE (6-мес runway)", "пример ~$1.5M", "Да (освобождается по мере продаж)", "Ставка/размер по числу закрытых сделок за 36 мес", "🟡 возможно, но условия хуже без exits", "Да", "(833) 315-5112", "limaone.com/line-of-credit", 2),
    ("Constructive Capital", "Fix&flip revolving LINE", "до $2M на проект", "Да (draw→продал/рефи→повтор)", "Для 'scaling' (2+ параллельных проекта)", "❌ нет (для масштабирующихся)", "Да", "форма на сайте", "constructiveloans.com/fix-flip", 2),
    ("RCN Capital", "Revolving line (под equity) + мелкая unsecured", "unsecured до $100k", "Да (6-мес, недельный ACH)", "По кредиту; лучшие условия — 2+ сделок/год", "🟡 линия под опыт; per-deal новичку", "Да (на займах)", "(860) 432-4782", "rcncapital.com", 2),
    ("Nectar", "Портфельное транш-финансирование (для BRRRR-держателей)", "$30k – $2M", "Частично (транши, не redraw-at-will)", "Нужен опыт управления арендой 3+ года", "❌ нет", "Да (по кэшфлоу портфеля)", "usenectar.com", 2),
    ("Anchor Loans", "Facility для builder/community + per-deal flip", "проектно, до ~$3M+", "Частично (facility для комьюнити-программ)", "Лучшие условия — 5+ флипов за 18 мес", "❌ нет", "Да", "anchorloans.com/solutions-for/renovators", 2),
    # НЕ линия (per-deal) — чтобы не гоняться
    ("BridgeWell Capital", "❌ НЕ линия — per-deal rehab holdback", "заём $100k–$2M (не линия)", "Нет (draws внутри одной сделки)", "FICO 620+, 'без опыта', FL (Orlando)", "✅ per-deal да (но это не линия)", "Да", "(866) 500-4500", "bridgewellcapital.com/rehab-only", 3),
    ("Kiavi", "❌ НЕ линия инвестору (per-deal)", "заём до $3M", "Нет (Enterprise = per-deal, 30+ объектов)", "Fix-flip 95% LTC, без дохода, LLC", "per-deal да", "Да", "(844) 415-4663", "kiavi.com", 3),
    ("Easy Street / Backflip / Roc360 / Renovo", "❌ НЕ линия — per-deal", "per-deal", "Нет", "—", "per-deal", "Да", "см. сайты", "easystreetcap.com · backflip.com", 3),
]
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Credit lines"
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="2E4A62")
thin = Side(style="thin", color="D0D0D0"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
ws.append(F)
for c in ws[1]: c.font = hf; c.fill = hfill; c.alignment = Alignment(vertical="center", wrap_text=True); c.border = bd
colf = {1: PatternFill("solid", fgColor="E7F0E9"), 2: PatternFill("solid", fgColor="FDF3D8"), 3: PatternFill("solid", fgColor="F3E9E9")}
rows.sort(key=lambda r: (r[-1], r[0]))
for r in rows:
    ws.append(list(r[:-1])); row = ws[ws.max_row]; rf = colf.get(r[-1])
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = bd
        if rf: c.fill = rf
for i, w in enumerate([26, 34, 22, 26, 40, 26, 18, 24, 40], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(F))}{ws.max_row}"
out = os.path.join(os.path.dirname(__file__), "..", "reports", "credit-lines.xlsx")
wb.save(out); print("wrote", out, "·", len(rows), "провайдеров")
