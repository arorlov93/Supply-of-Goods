#!/usr/bin/env python3
"""Чистый план обзвона под НОВУЮ компанию без опыта (клининг + субподряд, Флорида).
Оставляем только actionable FL-строки: GC final-clean + регистрация вендора + SBE-пул.
Выкидываем протухшие и out-of-state SAM-тендеры. + лист СКРИПТ с готовыми репликами."""
import csv, os, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = os.path.join(os.path.dirname(__file__), "..", "reports", "call-list-2026-08-25.csv")
OUT = os.path.join(os.path.dirname(__file__), "..", "reports", "call-plan-2026-08-26.xlsx")
rows = list(csv.DictReader(open(SRC, encoding="utf-8-sig")))

def g(r, key):
    for k in r:
        if key in k: return (r[k] or "").strip()
    return ""

# ---------- собираем actionable ----------
gc, reg, sbe = [], [], []
seen_phone = set()
for r in rows:
    name = g(r, "аименов")
    phone = g(r, "елефон")
    city = g(r, "ород")
    contact = g(r, "онтакт")
    val = g(r, "тоимость")
    src = g(r, "сточник")
    link = g(r, "сылк")
    desc = g(r, "писание")
    if name.startswith("GC"):
        if phone and phone in seen_phone:  # дедуп по телефону
            continue
        seen_phone.add(phone)
        who = name.replace("GC (final-clean суб): ", "").title()
        addr = desc.replace("Активный объект: ", "")
        gc.append([who, city, phone, contact, addr, val, "GC"])
    elif name.startswith("ЗВОНОК"):
        who = name.replace("ЗВОНОК: ", "")
        reg.append([who, city, phone, contact, desc, src, link, "REG"])
    elif "JLS" in name or "SBE" in name:
        who = name.replace("ТЕНДЕР: ", "")
        sbe.append([who, city, phone, contact, desc, src, link, "SBE"])

# ---------- рабочая книга ----------
wb = openpyxl.Workbook()

# ===== Лист 1: СКРИПТ =====
s = wb.active; s.title = "СКРИПТ"
BRAND = "2E4A62"
h1 = Font(bold=True, size=14, color="FFFFFF")
h2 = Font(bold=True, size=11, color=BRAND)
body = Font(size=11)
fillb = PatternFill("solid", fgColor=BRAND)
def put(cell, text, font=body, fill=None, wrap=True):
    s[cell] = text; s[cell].font = font
    s[cell].alignment = Alignment(wrap_text=wrap, vertical="top")
    if fill: s[cell].fill = fill
s.column_dimensions["A"].width = 118
put("A1", "СКРИПТ ЗВОНКА — ISP Group LLC (новая компания, клининг + субподряд, Флорида)", h1, fillb)
blocks = [
 ("ПОЗИЦИОНИРОВАНИЕ (держать в голове — мы новые, без истории)",
  "• Мы — субподрядчик по УБОРКЕ и мелкому отделочному субподряду в Майами-Дейд/Бровард.\n"
  "• Во Флориде на клининг ЛИЦЕНЗИЯ НЕ НУЖНА (self-perform) — работать можем легально с первого дня.\n"
  "• Мы застрахованы (General Liability), есть EIN/W-9, готовы прислать COI на объект.\n"
  "• Мы НОВЫЕ — и превращаем это в плюс: конкурентная цена, гибкий график, берёмся за небольшие объёмы,\n"
  "  выходим на объект быстро, лично на связи. Готовы стартовать с одного объекта и показать качество.\n"
  "• Цель звонка: НЕ продать сразу, а узнать ПОТРЕБНОСТЬ и попасть в список на расчёт."),
 ("A. GC — FINAL / POST-CONSTRUCTION CLEAN (звонок по активному объекту)",
  "«Здравствуйте, это [Имя] из ISP Group. Мы делаем финальную уборку после стройки — rough clean,\n"
  "final clean, мойка окон, вывоз мусора — как субподряд для генподрядчиков. Увидел, у вас активный\n"
  "объект на [адрес]. Хотел предложить нас на этап сдачи.\n"
  "Мы новая компания, поэтому даём конкурентную цену и гибкий график; полностью застрахованы,\n"
  "во Флориде на уборку лицензия не требуется.\n"
  "Подскажите, пожалуйста:»\n"
  "   1) Кто у вас отвечает за final clean по объекту — с кем лучше это обсудить?\n"
  "   2) Какой метраж (SF) и когда планируется сдача / turnover?\n"
  "   3) Уже выбрали клинера или ещё смотрите предложения?\n"
  "   4) Какой ориентир бюджета на финальную уборку по этому объекту?\n"
  "→ «Пришлю расчёт от площади в тот же день. Куда удобнее — email или текст?»\n"
  "ЦЕНОВОЙ ЯКОРЬ (себе): final clean $0.30–0.45/SF · вывоз мусора от $495/30 yd."),
 ("B. РЕГИСТРАЦИЯ ВЕНДОРА (город / округ / университет / больница)",
  "«Здравствуйте, ISP Group, клининговая компания из Авентуры. Хотим зарегистрироваться у вас\n"
  "как вендор, чтобы участвовать в бидах на janitorial / grounds / pressure washing.\n"
  "Подскажите:»\n"
  "   1) Как встать в ваш вендор-лист — какой портал и что заполнить?\n"
  "   2) Нужна ли для клининг-бидов SBE / local-сертификация?\n"
  "   3) Какие ближайшие клининг/grounds-контракты на горизонте и их ориентир бюджета?\n"
  "   4) Есть ли контракты, где допускают новых/небольших вендоров или субподряд?\n"
  "→ Зарегистрироваться на портале из колонки «Куда подать», поставить фильтр на janitorial NAICS 561720."),
 ("C. MIAMI-DADE JLS / SBE-ПУЛ (клининг-задания для малого бизнеса)",
  "«Хотим попасть в пул Janitorial & Landscaping Services для малого бизнеса.\n"
  "Что нужно для SBE-сертификации, какой размер типового task order, и когда следующий набор?»\n"
  "ВАЖНО: для SBE-G&S часто нужен Local Business Tax Receipt 1+ год — уточнить, т.к. LLC новая."),
 ("ОТРАБОТКА ВОЗРАЖЕНИЯ «А какой у вас опыт / объекты?»",
  "«Компания новая, но команда на объектах не первый год. Именно поэтому даём цену ниже рынка\n"
  "и максимальную гибкость. Предлагаю дать нам один объект / участок — с полной страховкой и\n"
  "гарантией переделки за наш счёт, если что-то не так. Рискуете минимально, а цену видите сразу.»"),
 ("ЧЕК-ЛИСТ ПЕРЕД ЗВОНКАМИ",
  "☐ W-9 готов  ☐ COI (сертификат страховки) под рукой  ☐ email для отправки расчёта\n"
  "☐ прайс-якоря записаны  ☐ таблица открыта на листе «ОБЗВОН»  ☐ ручка для заметок в конце строки"),
]
row = 3
for title, txt in blocks:
    put(f"A{row}", title, h2); row += 1
    put(f"A{row}", txt, body)
    s.row_dimensions[row].height = 15 * (txt.count("\n") + 2)
    row += 2

# ===== Лист 2: ОБЗВОН =====
w = wb.create_sheet("ОБЗВОН")
F = ["Кого / объект", "Что это", "Город", "Телефон", "Контакт / email",
     "ДЕЙСТВИЕ", "Что сказать / спросить (кратко)", "Куда подать / портал", "Ориентир суммы"]
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor=BRAND)
thin = Side(style="thin", color="D0D0D0"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
w.append(F)
for c in w[1]:
    c.font = hf; c.fill = hfill; c.border = bd
    c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")

green = PatternFill("solid", fgColor="E7F0E9")   # регистрация
yellow = PatternFill("solid", fgColor="FDF3D8")  # SBE-пул
blue = PatternFill("solid", fgColor="E6EEF5")    # GC

GC_ASK = "Скрипт A. Спросить: кто отвечает за final clean · метраж SF · дата сдачи · выбрали ли клинера · бюджет на уборку. Прислать расчёт от площади."
REG_ASK = "Скрипт B. Спросить: как встать в вендор-лист · нужна ли SBE-серт. · ближайшие janitorial/grounds биды и их бюджет · берут ли новых/суб."
SBE_ASK = "Скрипт C. Спросить: что нужно для SBE-серт. · размер task order · когда набор. Уточнить Local Business Tax Receipt (LLC новая)."

def emit(r, kind, fill):
    w.append(r); rr = w[w.max_row]
    for c in rr:
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = bd; c.fill = fill

# 1) Регистрация (быстрые «встать в базу»)
for who, city, phone, contact, desc, src, link, _ in reg:
    emit([who, "Регистрация вендора", city, phone or "—", contact or "—",
          "ЗАРЕГИСТРИРОВАТЬСЯ + звонок", REG_ASK, (link or src), "—"], "REG", green)
# 2) SBE-пул
for who, city, phone, contact, desc, src, link, _ in sbe:
    emit([who, "SBE-пул (клининг для МСБ)", city, phone or "—", contact or "—",
          "ПОДАТЬ НА СЕРТ. + звонок", SBE_ASK, (link or src), "task order"], "SBE", yellow)
# 3) GC final-clean (ядро обзвона), по городу
gc.sort(key=lambda x: (x[1], x[0]))
for who, city, phone, contact, addr, val, _ in gc:
    emit([who, "GC — final clean (суб)", city, phone or "—", contact or "—",
          "ЗВОНИТЬ (предложить уборку)", GC_ASK, "Прямой суб-квоут (email/тел)", val or "—"], "GC", blue)

widths = [30, 22, 16, 18, 26, 24, 52, 30, 20]
for i, wd in enumerate(widths, 1):
    w.column_dimensions[get_column_letter(i)].width = wd
w.freeze_panes = "A2"
w.auto_filter.ref = f"A1:{get_column_letter(len(F))}{w.max_row}"

wb.save(OUT)
print(f"wrote {OUT}")
print(f"Регистрация: {len(reg)} · SBE-пул: {len(sbe)} · GC final-clean: {len(gc)} · всего строк: {len(reg)+len(sbe)+len(gc)}")
