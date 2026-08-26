#!/usr/bin/env python3
"""Обзвон по свежим коммерческим ремонтам: Miami-Dade (RER) + Broward (Fort Lauderdale).
Фокус: office / restaurant / retail / clinic / salon / gym — прямой звонок GC и заказчику."""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SP = "/tmp/claude-0/-home-user-Supply-of-Goods/4f5e4e6d-222f-57da-93ad-94f5b2132f9f/scratchpad"

# целевые назначения (маленький офис/ресторан/ритейл/клиника/салон/зал)
TARGET = {
    "RETAIL SALES": "Ритейл", "OFFICE - PROFESSIONAL BUILDINGS": "Офис", "OFFICE USE ONLY": "Офис",
    "OFFICE - SALES": "Офис", "RESTAURANT-CAFETERIA": "Ресторан",
    "RESTAURANT/CAFET/BAR/LOUNGE/NIGHT CLUB": "Ресторан/бар", "BAR/COCKTAIL LOUNGE/RESTAURANTS": "Ресторан/бар",
    "BAR-LOUNGE-NIGHT CLUB": "Бар", "BAKERY PLANT": "Пекарня", "BANQUET HALL": "Банкетный зал",
    "CONVENIENCE STORE": "Магазин у дома", "PACKAGE STORE": "Ликёр-магазин",
    "CLINIC/SANITARIUMS/HEALTH CENTERS": "Клиника", "BEAUTY SALON-BARBER SHOP": "Салон/барбершоп",
    "GYM/EXERCISE CLUB": "Спортзал", "COIN LAUNDRY-DRY CLEANING": "Прачечная/химчистка",
    "DAYCARE - KINDERGARTEN": "Детсад",
}

def phone_fmt(p):
    d = "".join(c for c in (p or "") if c.isdigit())
    if len(d) < 10 or d[-10:] == "0000000000": return ""
    return f"({d[-10:-7]}) {d[-7:-4]}-{d[-4:]}"

def offer(use_ru, apptype, active):
    a = (apptype or "").upper()
    tag = "🟢 стройка идёт" if active else ""
    if use_ru in ("Ресторан", "Ресторан/бар", "Бар", "Пекарня", "Банкетный зал"):
        return f"Отделка-суб + FINAL CLEAN перед открытием {tag}".strip()
    if "NEW" in a or "SHELL" in a:
        return f"FINAL CLEAN (сдача) + отделка-суб {tag}".strip()
    return f"FINAL CLEAN после ремонта + отделка-суб {tag}".strip()

rows = []
seen = set()

# --- Miami-Dade ---
for off in (0, 1000):
    for a in [f["attributes"] for f in json.load(open(f"{SP}/mdc_{off}.json")).get("features", [])]:
        use = a.get("ProposedUseDescription")
        if use not in TARGET: continue
        ph = phone_fmt(a.get("ContractorPhone"))
        gc = (a.get("ContractorName") or "").strip()
        if not ph or not gc: continue
        key = (gc, ph, (a.get("PropertyAddress") or "").strip())
        if key in seen: continue
        seen.add(key)
        active = not a.get("CoCcDate")
        rows.append(["Miami-Dade", (a.get("OwnerName") or "").title(), gc.title(), ph,
                     (a.get("PropertyAddress") or "").title(), TARGET[use],
                     (a.get("ApplicationTypeDescription") or "").title().replace("Alter", "Ремонт"),
                     (a.get("PermitIssuedDate") or ""), "нет" if active else "да",
                     offer(TARGET[use], a.get("ApplicationTypeDescription"), active)])

# --- Broward (Fort Lauderdale) — из прошлого пула, только целевые/все коммерч. ремонты ---
def ftl_use(uc):
    u = (uc or "").upper()
    for k, ru in [("RESTAURANT","Ресторан"),("RETAIL","Ритейл"),("OFFICE","Офис"),("MEDICAL","Клиника"),
                  ("BAKERY","Пекарня"),("BAR","Бар"),("SPA","Салон"),("HEALTH","Клиника")]:
        if k in u: return ru
    return "Коммерция"
for fn in ("ftl3.json", "ftl4.json"):
    for a in [f["attributes"] for f in json.load(open(f"{SP}/{fn}")).get("features", [])]:
        ph = phone_fmt(a.get("CONTRACTPH"))
        gc = (a.get("CONTRACTOR") or "").strip()
        if not ph or not gc: continue
        key = (gc, ph, (a.get("FULLADDR") or "").strip())
        if key in seen: continue
        seen.add(key)
        use_ru = ftl_use(a.get("USECLASS"))
        rows.append(["Broward (FTL)", (a.get("OWNERNAME") or "").title() if a.get("OWNERNAME") else "",
                     gc.title(), ph, (a.get("FULLADDR") or "").title(), use_ru,
                     (a.get("PERMITDESC") or "").replace(" Permit", ""),
                     "", "", offer(use_ru, a.get("PERMITDESC"), True)])

# сорт: сначала целевые уборочные (ресторан/новый), потом по дате
def rank(r):
    tgt = 0 if r[5] in ("Ресторан","Ресторан/бар","Бар","Пекарня","Ритейл","Клиника","Салон/барбершоп","Спортзал") else 1
    return (tgt, r[7] == "", r[7])
rows.sort(key=lambda r: (rank(r)[0], r[7] == "", ), reverse=False)
rows.sort(key=lambda r: r[7], reverse=True)  # date desc primary among same

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Обзвон по пермитам"
BRAND = "2E4A62"
F = ["Округ", "Заказчик (владелец)", "Подрядчик (GC)", "Телефон GC", "Объект (адрес)",
     "Назначение", "Тип работ", "Дата пермита", "CO выдан?", "Что предложить в звонке"]
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor=BRAND)
thin = Side(style="thin", color="D0D0D0"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
ws.append(F)
for c in ws[1]:
    c.font = hf; c.fill = hfill; c.border = bd
    c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
food = PatternFill("solid", fgColor="FDE9D8"); retail = PatternFill("solid", fgColor="E7F0E9")
for r in rows:
    ws.append(r); row = ws[ws.max_row]
    fill = food if r[5] in ("Ресторан","Ресторан/бар","Бар","Пекарня","Банкетный зал") else (
           retail if r[5] in ("Ритейл","Магазин у дома","Клиника","Салон/барбершоп","Спортзал") else None)
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = bd
        if fill: c.fill = fill
for i, wd in enumerate([13, 24, 26, 15, 26, 16, 18, 13, 10, 38], 1):
    ws.column_dimensions[get_column_letter(i)].width = wd
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(F))}{ws.max_row}"
out = os.path.join(os.path.dirname(__file__), "..", "reports", "permit-callist-2026-08-26.xlsx")
wb.save(out)
md = sum(1 for r in rows if r[0]=="Miami-Dade"); bw = len(rows)-md
print(f"wrote {out} · всего {len(rows)} (Miami-Dade {md}, Broward {bw})")
