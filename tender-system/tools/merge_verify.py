#!/usr/bin/env python3
"""Свести веб-сверку телефонов (батчи A/B/C) в приоритетный файл.
Добавляет колонки: Веб-сверка (✓ совпал / ≠ другой № / — не проверен) + Альт. телефон (веб)."""
import json, os, re, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SP = "/tmp/claude-0/-home-user-Supply-of-Goods/4f5e4e6d-222f-57da-93ad-94f5b2132f9f/scratchpad"
verify = json.load(open(f"{SP}/verify_p1.json"))  # index -> {gc,phone,addr,date}

# parse batch txts: index | STATUS | county:.. | found:.. | source:..
res = {}
for fn in ("verify_A.txt", "verify_B.txt", "verify_C.txt"):
    for ln in open(f"{SP}/{fn}"):
        ln = ln.strip()
        if not ln or "|" not in ln: continue
        parts = [p.strip() for p in ln.split("|")]
        try: idx = int(parts[0])
        except: continue
        status = parts[1]
        found = ""
        for p in parts:
            if p.startswith("found:"):
                found = p[6:].strip()
        if found == "-": found = ""
        res[idx] = (status, found)

# map (gc,phone) -> (status, found)
vmap = {}
for i, v in enumerate(verify):
    if i in res:
        vmap[(v["gc"], v["phone"])] = res[i]

STAT = {"MATCH": "✓ совпал (веб)", "DIFFERENT": "≠ другой № на сайте", "NOT_FOUND": "— не найден в вебе"}

# regenerate priority rows (same logic as make_permit_priority)
TODAY = datetime.date(2026, 8, 26)
TARGET = {
    "RETAIL SALES": "Ритейл", "OFFICE - PROFESSIONAL BUILDINGS": "Офис", "OFFICE USE ONLY": "Офис",
    "OFFICE - SALES": "Офис", "RESTAURANT-CAFETERIA": "Ресторан",
    "RESTAURANT/CAFET/BAR/LOUNGE/NIGHT CLUB": "Ресторан/бар", "BAR/COCKTAIL LOUNGE/RESTAURANTS": "Ресторан/бар",
    "BAR-LOUNGE-NIGHT CLUB": "Бар", "BAKERY PLANT": "Пекарня", "BANQUET HALL": "Банкетный зал",
    "CONVENIENCE STORE": "Магазин у дома", "PACKAGE STORE": "Ликёр-магазин",
    "CLINIC/SANITARIUMS/HEALTH CENTERS": "Клиника", "BEAUTY SALON-BARBER SHOP": "Салон/барбершоп",
    "GYM/EXERCISE CLUB": "Спортзал", "COIN LAUNDRY-DRY CLEANING": "Прачечная/химчистка",
    "DAYCARE - KINDERGARTEN": "Детсад",
}
def d2(s):
    try: return datetime.date.fromisoformat(s[:10])
    except: return None
def phone_fmt(p):
    d = "".join(c for c in (p or "") if c.isdigit())
    if len(d) < 10 or d[-10:] == "0000000000": return ""
    return f"({d[-10:-7]}) {d[-7:-4]}-{d[-4:]}"
def classify(app, cmt):
    a = (app or "").upper(); c = (cmt or "").upper()
    demo = bool(re.search(r"\bDEMO\b|DEMOLISH|DEMOLITION", c)) and "NEW" not in c
    interior = ("ALTER - INTERIOR" in a) or any(k in c for k in
        ("REMODEL","BUILDOUT","BUILD OUT","RENO","TENANT","ALTERATION","ALT INT","INT ALT","CHANGE OF USE","CHANGE USE","LEGALIZATION","ESTABLISH USE","FINISH"))
    newb = any(k in a for k in ("NEW","SHELL","ADDITION"))
    if demo and not interior and not newb: return 3, "ДЕМОНТАЖ", "Демонтаж/вывоз"
    if interior: return 1, "ОТДЕЛКА", "Покраска + гипсокартон + полы"
    if newb: return 2, "УБОРКА", "Final clean на сдаче (+ отделка позже)"
    return 2, "УБОРКА", "Final clean после работ"

feats = []
for off in (0, 1000):
    feats += [f["attributes"] for f in json.load(open(f"{SP}/md_{off}.json")).get("features", [])]
rows = []; seen = set()
for a in feats:
    use = a.get("ProposedUseDescription")
    if use not in TARGET: continue
    ph = phone_fmt(a.get("ContractorPhone")); gc = (a.get("ContractorName") or "").strip()
    dt = d2(a.get("PermitIssuedDate"))
    if not ph or not gc or not dt: continue
    key = (gc, ph, (a.get("PropertyAddress") or "").strip())
    if key in seen: continue
    seen.add(key)
    age = (TODAY - dt).days
    tier = "🟢 ≤30д" if age <= 30 else ("🟢 31-90д" if age <= 90 else "🟡 91-180д")
    pri, ptag, svc = classify(a.get("ApplicationTypeDescription"), a.get("DetailDescriptionComments"))
    st, found = vmap.get((gc.title(), ph), ("", ""))
    rows.append([pri, ptag, svc, (a.get("OwnerName") or "").title(), gc.title(), ph,
                 (a.get("PropertyAddress") or "").title(), TARGET[use],
                 (a.get("ApplicationTypeDescription") or "").title().replace("Alter", "Ремонт"),
                 (a.get("DetailDescriptionComments") or "").strip().title(),
                 dt.isoformat(), tier, STAT.get(st, ""), found])
rows.sort(key=lambda r: (r[0], r[10] and (TODAY - datetime.date.fromisoformat(r[10])).days))

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Обзвон (сверено)"
BRAND = "2E4A62"
F = ["Приоритет", "Услуга", "Заказчик", "Подрядчик (GC)", "Телефон GC (из пермита)", "Объект",
     "Назначение", "Тип работ", "Описание", "Дата", "Актуальность", "Веб-сверка", "Альт. телефон (веб)"]
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor=BRAND)
thin = Side(style="thin", color="D0D0D0"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
ws.append([F[0], F[1], F[2], F[3], F[4], F[5], F[6], F[7], F[8], F[9], F[10], F[11], F[12]])
for c in ws[1]:
    c.font = hf; c.fill = hfill; c.border = bd
    c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
pcol = {1: PatternFill("solid", fgColor="DCE9F7"), 2: PatternFill("solid", fgColor="FDE9D8"), 3: PatternFill("solid", fgColor="ECECEC")}
green = Font(color="1E7A34", bold=True); amber = Font(color="9A6A00"); grey = Font(color="777777")
for r in rows:
    ws.append([f"P{r[0]} {r[1]}", r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[9], r[10], r[11], r[12], r[13]])
    row = ws[ws.max_row]
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = bd; c.fill = pcol[r[0]]
    vc = row[11]
    if r[12].startswith("✓"): vc.font = green
    elif r[12].startswith("≠"): vc.font = amber
    elif r[12].startswith("—"): vc.font = grey
for i, wd in enumerate([13, 26, 20, 24, 18, 22, 14, 15, 18, 12, 12, 20, 18], 1):
    ws.column_dimensions[get_column_letter(i)].width = wd
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(F))}{ws.max_row}"
out = os.path.join(os.path.dirname(__file__), "..", "reports", "permit-priority-VERIFIED-2026-08-26.xlsx")
wb.save(out)
from collections import Counter
cc = Counter(x[0] for x in res.values())
print("wrote", out)
print("сверено (топ-45 P1):", dict(cc))
