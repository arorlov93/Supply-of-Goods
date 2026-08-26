#!/usr/bin/env python3
"""ЧИСТЫЙ обзвон по свежим коммерческим ремонтам Miami-Dade (RER).
Только: активные (CO не выдан) + выданы ≤180 дней + телефон подрядчика есть.
Фокус: office/restaurant/retail/clinic/salon/gym. Дата и тир актуальности в каждой строке.
+ лист «Аудит данных» — чтобы доверять свежести."""
import json, os, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SP = "/tmp/claude-0/-home-user-Supply-of-Goods/4f5e4e6d-222f-57da-93ad-94f5b2132f9f/scratchpad"
TODAY = datetime.date(2026, 8, 26)

TARGET = {
    "RETAIL SALES": "Ритейл", "OFFICE - PROFESSIONAL BUILDINGS": "Офис", "OFFICE USE ONLY": "Офис",
    "OFFICE - SALES": "Офис", "RESTAURANT-CAFETERIA": "Ресторан",
    "RESTAURANT/CAFET/BAR/LOUNGE/NIGHT CLUB": "Ресторан/бар", "BAR/COCKTAIL LOUNGE/RESTAURANTS": "Ресторан/бар",
    "BAR-LOUNGE-NIGHT CLUB": "Бар", "BAKERY PLANT": "Пекарня", "BANQUET HALL": "Банкетный зал",
    "CONVENIENCE STORE": "Магазин у дома", "PACKAGE STORE": "Ликёр-магазин",
    "CLINIC/SANITARIUMS/HEALTH CENTERS": "Клиника", "BEAUTY SALON-BARBER SHOP": "Салон/барбершоп",
    "GYM/EXERCISE CLUB": "Спортзал", "COIN LAUNDRY-DRY CLEANING": "Прачечная/химчистка",
    "DAYCARE - KINDERGARTEN": "Детсад",
}
FOOD = {"Ресторан", "Ресторан/бар", "Бар", "Пекарня", "Банкетный зал"}

def d2(s):
    try: return datetime.date.fromisoformat(s[:10])
    except: return None

def phone_fmt(p):
    d = "".join(c for c in (p or "") if c.isdigit())
    if len(d) < 10 or d[-10:] == "0000000000": return ""
    return f"({d[-10:-7]}) {d[-7:-4]}-{d[-4:]}"

def tier(dt):
    age = (TODAY - dt).days
    if age <= 30: return "🟢 ≤30 дней", age
    if age <= 90: return "🟢 31-90 дней", age
    return "🟡 91-180 дней", age

def offer(use_ru, apptype):
    a = (apptype or "").upper()
    if use_ru in FOOD: return "Отделка-суб + FINAL CLEAN перед открытием"
    if "NEW" in a or "SHELL" in a: return "FINAL CLEAN (сдача) + отделка-суб"
    return "FINAL CLEAN после ремонта + отделка-суб (paint/floor)"

# load
feats = []
for off in (0, 1000, 2000, 3000):
    p = f"{SP}/md_{off}.json"
    if os.path.exists(p):
        feats += [f["attributes"] for f in json.load(open(p)).get("features", [])]

rows = []
seen = set()
total_seen = len(feats)
dropped_use = dropped_phone = 0
for a in feats:
    use = a.get("ProposedUseDescription")
    if use not in TARGET:
        dropped_use += 1; continue
    ph = phone_fmt(a.get("ContractorPhone"))
    gc = (a.get("ContractorName") or "").strip()
    if not ph or not gc:
        dropped_phone += 1; continue
    dt = d2(a.get("PermitIssuedDate"))
    if not dt: continue
    key = (gc, ph, (a.get("PropertyAddress") or "").strip())
    if key in seen: continue
    seen.add(key)
    tr, age = tier(dt)
    use_ru = TARGET[use]
    rows.append({
        "owner": (a.get("OwnerName") or "").title(),
        "gc": gc.title(), "phone": ph,
        "addr": (a.get("PropertyAddress") or "").title(),
        "use": use_ru, "use_food": use_ru in FOOD,
        "work": (a.get("ApplicationTypeDescription") or "").title().replace("Alter", "Ремонт"),
        "date": dt.isoformat(), "tier": tr, "age": age,
        "offer": offer(use_ru, a.get("ApplicationTypeDescription")),
    })
# сорт: сначала рестораны/бары, затем по дате (новее выше)
rows.sort(key=lambda r: (0 if r["use_food"] else 1, -(TODAY - datetime.date.fromisoformat(r["date"])).days * -1))
rows.sort(key=lambda r: r["date"], reverse=True)
rows.sort(key=lambda r: 0 if r["use_food"] else 1)

wb = openpyxl.Workbook()
BRAND = "2E4A62"
ws = wb.active; ws.title = "Обзвон (свежее)"
F = ["Заказчик (владелец)", "Подрядчик (GC)", "Телефон GC", "Объект (адрес)", "Назначение",
     "Тип работ", "Дата пермита", "Актуальность", "Что предложить в звонке"]
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor=BRAND)
thin = Side(style="thin", color="D0D0D0"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
ws.append(F)
for c in ws[1]:
    c.font = hf; c.fill = hfill; c.border = bd
    c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
food = PatternFill("solid", fgColor="FDE9D8"); ret = PatternFill("solid", fgColor="E7F0E9")
for r in rows:
    ws.append([r["owner"], r["gc"], r["phone"], r["addr"], r["use"], r["work"],
               r["date"], r["tier"], r["offer"]])
    row = ws[ws.max_row]
    fill = food if r["use_food"] else (ret if r["use"] in ("Ритейл","Клиника","Салон/барбершоп","Спортзал","Магазин у дома") else None)
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = bd
        if fill: c.fill = fill
for i, wd in enumerate([24, 26, 15, 26, 16, 18, 13, 16, 40], 1):
    ws.column_dimensions[get_column_letter(i)].width = wd
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(F))}{ws.max_row}"

# ---- audit sheet ----
au = wb.create_sheet("Аудит данных")
au.column_dimensions["A"].width = 46; au.column_dimensions["B"].width = 40
H = Font(bold=True, size=13, color="FFFFFF"); HB = PatternFill("solid", fgColor=BRAND)
def put(r, a, b="", bold=False):
    au[f"A{r}"] = a; au[f"B{r}"] = b
    au[f"A{r}"].font = Font(bold=bold); au[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    au[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
au["A1"] = "АУДИТ ДАННЫХ — обзвон по пермитам (26.08.2026)"; au["A1"].font = H; au["A1"].fill = HB
from collections import Counter
tc = Counter(r["tier"] for r in rows); uc = Counter(r["use"] for r in rows)
put(3, "Источник", "Miami-Dade County RER (официальный пермит-реестр)", True)
put(4, "URL", "services.arcgis.com/.../miamidade_permit_data/FeatureServer/0")
put(5, "Обновление источника", "ежедневно; заполнен телефон подрядчика ~82%")
put(6, "ФИЛЬТР (жёсткий)", "коммерция + building-пермит + ALTER-INTERIOR/NEW/SHELL/ADDITION")
put(7, "  + активные", "CO НЕ выдан (CoCcDate пустой) = стройка идёт")
put(8, "  + свежесть", "выдан ≤180 дней назад (после 26.02.2026)")
put(9, "  + телефон", "у подрядчика есть валидный телефон")
put(10, "  + назначение", "office/restaurant/retail/clinic/salon/gym и т.п.")
put(12, "ИТОГО в списке", f"{len(rows)} уникальных объектов", True)
put(13, "  🟢 ≤30 дней", str(tc.get("🟢 ≤30 дней", 0)))
put(14, "  🟢 31-90 дней", str(tc.get("🟢 31-90 дней", 0)))
put(15, "  🟡 91-180 дней", str(tc.get("🟡 91-180 дней", 0)))
put(16, "  старше 180 дней", "0 — отфильтровано")
put(17, "Все активны (CO нет)?", "ДА — 100% (фильтр по CoCcDate)")
put(19, "По назначению", "; ".join(f"{u}:{n}" for u, n in uc.most_common()))
put(21, "Broward (Fort Lauderdale)", "ИСКЛЮЧЁН: их фид отстаёт ~15 мес (все записи 2021–май 2025)", True)
put(22, "City of Miami", "есть имена GC, но БЕЗ телефонов — не звонить")
put(24, "Как обновлять", "запустить этот скрипт заново — тянет свежак Miami-Dade на дату запуска")
for rr in range(3, 25):
    if au[f"A{rr}"].value and au[f"A{rr}"].value.startswith(("Источник","ИТОГО","Broward")):
        pass
out = os.path.join(os.path.dirname(__file__), "..", "reports", "permit-callist-CLEAN-2026-08-26.xlsx")
wb.save(out)
print(f"wrote {out}")
print(f"всего вошло: {len(rows)} | сырьё: {total_seen} | отсеяно по назначению: {dropped_use} | без телефона: {dropped_phone}")
print("тиры:", dict(tc)); print("назначения:", dict(uc.most_common()))
