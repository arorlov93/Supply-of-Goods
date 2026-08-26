#!/usr/bin/env python3
"""Excel: кредиторы под аукционы (hard-money/bridge/DSCR), приоритет Флорида."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F = ["Компания", "FL?", "Продукты", "Аукцион / быстрое закрытие", "Линия / POF", "LLC + без дохода", "Телефон", "Email", "Сайт"]
# tier: 1=FL auction-friendly, 2=national, 3=DSCR-only (для рефинанса, не для аукциона)
rows = [
    ("EquityMax", "FL (Ft Lauderdale)", "Hard money fix&flip, bridge, cash-out, land", "⭐ Спец. страница 'auction 48h'; POF за ~5 мин, безлимитно", "POF (не линия)", "Да — asset-based, без дохода", "(954) 267-9103 / (954) 771-2407", "info@equitymax.com", "equitymax.com/auction-property-loans", 1),
    ("Gauntlet Funding", "NY (проверить FL!)", "Hard money, fix&flip (REO/foreclosure)", "⭐ Явный POF под аукцион; закрытие 1–2 дня", "POF (не линия)", "Да — минимум документов", "см. сайт", "", "gauntletfunding.com/proof-of-funds", 1),
    ("BridgeWell Capital", "FL (Orlando, по штату)", "Fix-to-flip, rehab, cash-out, rental", "Закрытие от 10 дней", "⭐ ЕДИНСТВ. с 'rehab CREDIT LINE'", "Да — плохой кредит/self-employed ок", "(407) 447-5000 / (866) 500-4500", "", "bridgewellcapital.com", 1),
    ("DKC Lending", "FL (Tampa)", "Hard money fix&flip (70% LTV, 50% ремонта)", "'Cash offer' закрытие за 3 дня", "POF (не линия)", "Да — asset-based", "(813) 501-5729", "info@dkclending.com", "dkclending.com", 1),
    ("Titan Funding", "FL (Boca Raton)", "Hard money, bridge, fix&flip, ground-up", "Решение 48ч, фандинг за 4 дня ($100K–5M)", "Per-deal", "Да — частный/asset", "(855) 731-1600 / (561) 756-9683", "info@titanfunding.com", "titanfunding.com", 1),
    ("M&M Private Lending", "FL (Ft Lauderdale)", "Hard money res/comm", "Быстрые asset-based закрытия", "Per-deal", "Да — без tax returns/credit/bank stmt", "(305) 899-2201", "loans@mmprivatelending.com", "mmprivatelending.com", 1),
    ("Vaster", "FL (Miami)", "Res/comm bridge, private money", "'В днях, не неделях'; до ~65% value", "Per-deal", "Да (hard-money опция)", "форма на сайте", "", "vaster.com", 1),
    ("LendingOne", "FL (Boca Raton), нац.", "Fix&flip, fix-to-rent, DSCR, bridge, new constr, portfolio", "Быстрое закрытие; POF пред-одобренным", "Portfolio-программы", "Да — LLC, DSCR/asset", "(866) 730-4032", "", "lendingone.com", 1),
    ("Kiavi", "Нац. (актив. FL)", "Fix&flip, bridge, DSCR, new constr, portfolio", "Быстрые tech-закрытия; POF в портале", "Repeat-borrower пайплайн", "Да — LLC, FICO~650", "(844) 415-4663", "", "kiavi.com/florida-fix-and-flip-loans", 2),
    ("RCN Capital", "Нац. (актив. FL)", "Fix&flip/bridge, DSCR, ground-up", "Закрытие от 10 раб. дней; 7–14 дн pre-approval→close", "Portfolio-опции", "Да — entity, collateral", "(860) 787-8262", "", "rcncapital.com/loan-programs", 2),
    ("Lima One Capital", "Нац. (актив. FL)", "Fix&flip, bridge, DSCR, new constr, multifamily", "POF-письмо; закрытия ~8 раб. дней", "Per-deal", "Да — entity", "(800) 390-4212", "", "limaone.com/hard-money-fix-n-flip", 2),
    ("CIVIC Financial (Roc360)", "Нац. (актив. FL)", "Fix&flip bridge (90% LTP), DSCR", "⭐ Закрытие от 5 раб. дней; BPO вместо оценки", "Институц. капитал", "Да — нет мин. FICO, по активу", "форма на сайте", "", "civicfs.com/fix-and-flip", 2),
    ("Roc Capital / Roc360", "Нац. (актив. FL)", "Fix&flip, ground-up, multifamily, DSCR ($75K–25M)", "Стримлайн, быстрое закрытие", "Table-funding/пайплайн", "Да — asset-based", "форма на сайте", "", "roccapital.com/fix-and-flip", 2),
    ("Easy Street Capital", "Нац. (стр. FL)", "Fix&flip (93% LTC), bridge, new constr, DSCR", "Быстрые закрытия; POF на контакт-стр.", "Per-deal", "Да — DSCR, без мин. DSCR опция", "форма на сайте", "", "easystreetcap.com/hard-money-loan-florida", 2),
    ("Anchor Loans", "Нац. (актив. FL)", "Fix&flip/bridge (95% LTC, $50K–10M)", "Fintech-ускоренный фандинг", "Per-deal (для опытных)", "Да — entity, asset/опыт", "(888) 719-5299", "", "anchorloans.com/loans/fix-and-flip-loans", 2),
    ("Constructive Loans", "Нац. wholesale (FL)", "Rehab fix&flip, DSCR (SFR/2-4/5-8)", "~20 дней (через брокера — медленнее)", "Per-deal", "Да — business-purpose", "форма на сайте", "", "constructiveloans.com/fix-flip", 2),
    ("Temple View Capital", "Нац. (актив. FL)", "Fix&flip, bridge, DSCR, new constr", "Быстрое финансирование (без аукц-стр.)", "Per-deal", "Да — asset-based", "(844) 900-3828", "", "templeviewcap.com", 2),
    ("Renovo Financial", "Нац. (офисы Miami/Orlando/Tampa)", "Fix&flip, ground-up, DSCR, multifamily, homebuilder", "Локальные офисы для скорости", "Per-deal", "Да — asset-based", "форма на сайте", "", "renovofinancial.com", 2),
    ("A&D Mortgage", "FL (Hollywood), нац.", "DSCR (30-yr, no income/employment), Non-QM", "НЕ аукцион — для рефинанса/hold", "DSCR-рефи", "Да — по арендному доходу", "форма на сайте", "", "admortgage.com/programs/dscr", 3),
    ("Visio Lending", "Нац. (актив. FL)", "DSCR long-term + STR/vacation", "НЕ аукцион — для рефинанса/hold (#1 DSCR SFR)", "DSCR-рефи", "Да — по аренде, entity", "(888) 521-0353", "loans@visiolending.com", "visiolending.com", 3),
]
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Auction lenders"
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="2E4A62")
thin = Side(style="thin", color="D0D0D0"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
ws.append(F)
for c in ws[1]: c.font = hf; c.fill = hfill; c.alignment = Alignment(vertical="center", wrap_text=True); c.border = bd
colf = {1: PatternFill("solid", fgColor="E7F0E9"), 2: None, 3: PatternFill("solid", fgColor="F3E9E9")}
rows.sort(key=lambda r: (r[-1], r[0]))
for r in rows:
    ws.append(list(r[:-1])); row = ws[ws.max_row]; rf = colf.get(r[-1])
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True); c.border = bd
        if rf: c.fill = rf
for i, w in enumerate([22, 20, 34, 30, 22, 24, 24, 24, 34], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(F))}{ws.max_row}"
import os
out = os.path.join(os.path.dirname(__file__), "..", "reports", "auction-lenders.xlsx")
wb.save(out); print("wrote", out, "·", len(rows), "кредиторов")
