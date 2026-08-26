#!/usr/bin/env python3
"""ИТОГОВЫЙ документ обзвона — 3 вкладки: P1 Отделка · P2 Уборка · P3 Демонтаж.
Описание по-русски + оценка бюджета НАШЕЙ работы от площади объекта.
Miami-Dade RER, активные (CO нет), свежесть ≤180д, телефон подрядчика. + веб-сверка топ-45 P1."""
import json, os, re, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SP = "/tmp/claude-0/-home-user-Supply-of-Goods/4f5e4e6d-222f-57da-93ad-94f5b2132f9f/scratchpad"
TODAY = datetime.date(2026, 8, 26)
TARGET = {
    "RETAIL SALES": "Ритейл", "OFFICE - PROFESSIONAL BUILDINGS": "Офис", "OFFICE USE ONLY": "Офис",
    "OFFICE - SALES": "Офис", "RESTAURANT-CAFETERIA": "Ресторан",
    "RESTAURANT/CAFET/BAR/LOUNGE/NIGHT CLUB": "Ресторан/бар", "BAR/COCKTAIL LOUNGE/RESTAURANTS": "Ресторан/бар",
    "BAR-LOUNGE-NIGHT CLUB": "Бар", "BAKERY PLANT": "Пекарня", "BANQUET HALL": "Банкетный зал",
    "CONVENIENCE STORE": "Магазин у дома", "PACKAGE STORE": "Ликёр-магазин",
    "CLINIC/SANITARIUMS/HEALTH CENTERS": "Клиника", "BEAUTY SALON-BARBER SHOP": "Салон/барбершоп",
    "GYM/EXERCISE CLUB": "Спортзал", "COIN LAUNDRY-DRY CLEANING": "Прачечная/химчистка",
    "DAYCARE - KINDERGARTEN": "Детсад", "WAREHOUSE/STORAGE": "Склад", "WAREHOUSE": "Склад",
}
# ВЕРХНЕ-СРЕДНЯЯ рыночная ставка $/SF (Miami-Dade/Broward, 2025-26, из research): цена, которую платит GC субу
RATE = {"P1": 22.0, "P2": 0.55, "P3": 7.0}
MINJOB = {"P2": 1500.0}          # мин. цена работы (final clean коммерч. ~ от $1500)
PROFIT = 0.20                    # целевая прибыль пользователя
# доля от стоимости стройки, если метраж не указан (верхне-средняя)
FRAC = {"P1": 0.22, "P2": 0.015, "P3": 0.08}

def d2(s):
    try: return datetime.date.fromisoformat(s[:10])
    except: return None
def phone_fmt(p):
    d = "".join(c for c in (p or "") if c.isdigit())
    if len(d) < 10 or d[-10:] == "0000000000": return ""
    return f"({d[-10:-7]}) {d[-7:-4]}-{d[-4:]}"
def tier(age):
    return "🟢 ≤30д" if age <= 30 else ("🟢 31-90д" if age <= 90 else "🟡 91-180д")
def num(s):
    try: return float(re.sub(r"[^0-9.]", "", str(s or "")))
    except: return 0.0

# --- перевод описания на русский ---
DESC_MAP = [
    (("TENANT IMPROVEMENT", "TENANT"), "Отделка под арендатора (build-out)"),
    (("INTERIOR BUILDOUT", "INTERIOR BUILD OUT", "BUILDOUT", "BUILD OUT"), "Отделка помещения под ключ (build-out)"),
    (("CHANGE OF USE", "CHANGE USE", "ESTABLISH USE"), "Смена назначения помещения (полный ремонт)"),
    (("INTERIOR RENOVATION", "INTERIOR RENO", "INT RENO", "RENOVATION", "RENO"), "Внутренняя реновация"),
    (("INTERIOR ALTERATION", "ALTERATION INTERIOR", "INT ALTERATION", "INTERIOR ALT", "ALT INT", "ALTERATION"), "Внутренняя перепланировка"),
    (("INTERIOR REMODELING", "INTERIOR REMODEL", "INT REMODEL", "INT REMOD", "REMODELING/REPAIRS", "REMODELING", "REMODEL"), "Внутренний ремонт помещения"),
    (("NEW CONSTRUC", "NEW CONSTRUCTION", "NEW STORE", "NEW WAREHOUSE", "NEW BLDG", "GROUND UP", "NEW "), "Новое строительство"),
    (("SHELL",), "Возведение коробки (shell)"),
    (("ADDITION",), "Пристройка"),
    (("LEGALIZATION", "LEGALIZE"), "Легализация ранее выполненных работ"),
    (("WALK IN COOLER", "COOLER"), "Установка холодильной камеры"),
    (("WALL OPENING", "WALL OPENINGS"), "Устройство проёма в стене"),
    (("PARTIAL DEMO", "DEMOLITION", "DEMOLISH", "DEMO"), "Демонтаж / снос"),
    (("STORAGE/OFFICE", "STORAGE"), "Складско-офисные работы"),
    (("REPAIR",), "Ремонт/восстановление"),
]
def ru_desc(cmt, app):
    txt = (cmt or "").upper() + " " + (app or "").upper()
    for keys, ru in DESC_MAP:
        if any(k in txt for k in keys):
            return ru
    return "Коммерческий ремонт"
APP_RU = {"ALTER - INTERIOR": "Внутр. ремонт", "NEW": "Новое стр-во", "SHELL ONLY": "Коробка (shell)",
          "ADDITION - ATTACHED": "Пристройка", "DEMOLISH": "Снос", "ALTER - EXTERIOR": "Наружн. ремонт",
          "REPAIR": "Ремонт"}

def budget(pri, sf, val):
    approx = ""
    if sf and sf > 50:
        b = sf * RATE[pri]; sfs = f"{sf:,.0f} SF"
    elif val and val > 500:
        b = val * FRAC[pri]; sfs = "метраж не указан"; approx = "≈"
    else:
        return "уточнить у GC", "", "", (f"{sf:,.0f} SF" if sf else "—")
    b = max(b, MINJOB.get(pri, 0))
    prof = b * PROFIT; cap = b * (1 - PROFIT)
    return f"{approx}${round(b,-2):,.0f}", f"${round(prof,-2):,.0f}", f"${round(cap,-2):,.0f}", sfs

# --- веб-сверка (топ-45 P1) ---
verify = json.load(open(f"{SP}/verify_p1.json"))
res = {}
for fn in ("verify_A.txt", "verify_B.txt", "verify_C.txt"):
    for ln in open(f"{SP}/{fn}"):
        parts = [p.strip() for p in ln.strip().split("|")]
        if len(parts) < 2 or not parts[0].isdigit(): continue
        found = next((p[6:].strip() for p in parts if p.startswith("found:")), "")
        res[int(parts[0])] = (parts[1], "" if found == "-" else found)
STAT = {"MATCH": "✓ совпал", "DIFFERENT": "≠ другой №", "NOT_FOUND": "— нет в вебе"}
vmap = {}
for i, v in enumerate(verify):
    if i in res:
        st, found = res[i]
        vmap[(v["gc"], v["phone"])] = (STAT.get(st, ""), found)

def rowdata(a, pri, svc):
    ph = phone_fmt(a.get("ContractorPhone")); gc = (a.get("ContractorName") or "").strip()
    dt = d2(a.get("PermitIssuedDate"))
    if not ph or not gc or not dt: return None
    age = (TODAY - dt).days
    st, found = vmap.get((gc.title(), ph), ("", ""))
    sf = num(a.get("SquareFootage")); val = num(a.get("EstimatedValue"))
    bud, prof, cap, sfs = budget(pri, sf, val)
    return {"owner": (a.get("OwnerName") or "").title(), "gc": gc.title(), "phone": ph,
            "addr": (a.get("PropertyAddress") or "").title(),
            "use": TARGET.get(a.get("ProposedUseDescription"), (a.get("ProposedUseDescription") or "").title()),
            "work": APP_RU.get((a.get("ApplicationTypeDescription") or "").upper(), (a.get("ApplicationTypeDescription") or "").title()),
            "desc": ru_desc(a.get("DetailDescriptionComments"), a.get("ApplicationTypeDescription")),
            "sf": sfs, "bud": bud, "prof": prof, "cap": cap, "date": dt.isoformat(), "tier": tier(age),
            "st": st, "found": found, "svc": svc, "age": age}

# --- P1/P2 ---
feats = []
for off in (0, 1000):
    feats += [f["attributes"] for f in json.load(open(f"{SP}/md2_{off}.json")).get("features", [])]
p1 = []; p2 = []; seen = set()
for a in feats:
    if a.get("ProposedUseDescription") not in TARGET: continue
    app = (a.get("ApplicationTypeDescription") or "").upper(); cmt = (a.get("DetailDescriptionComments") or "").upper()
    key = ((a.get("ContractorName") or "").strip(), phone_fmt(a.get("ContractorPhone")), (a.get("PropertyAddress") or "").strip())
    if not all(key) or key in seen: continue
    interior = ("ALTER - INTERIOR" in app) or any(k in cmt for k in ("REMODEL","BUILDOUT","BUILD OUT","RENO","TENANT","ALTERATION","ALT INT","INT ALT","CHANGE OF USE","CHANGE USE","LEGALIZATION","ESTABLISH USE","FINISH"))
    newb = any(k in app for k in ("NEW", "SHELL", "ADDITION"))
    if interior:
        r = rowdata(a, "P1", "Покраска + гипсокартон + полы")
        if r: p1.append(r); seen.add(key)
    elif newb:
        r = rowdata(a, "P2", "Final clean на сдаче (+ отделка позже)")
        if r: p2.append(r); seen.add(key)
# --- P3 ---
p3 = []; seen3 = set()
for a in [x["attributes"] for x in json.load(open(f"{SP}/demo2.json")).get("features", [])]:
    if a.get("ProposedUseDescription") not in TARGET: continue
    if "DEMO" not in (a.get("ApplicationTypeDescription") or "").upper(): continue
    key = ((a.get("ContractorName") or "").strip(), phone_fmt(a.get("ContractorPhone")), (a.get("PropertyAddress") or "").strip())
    if not all(key) or key in seen3: continue
    r = rowdata(a, "P3", "Демонтаж / вывоз мусора (суб)")
    if r: p3.append(r); seen3.add(key)
for lst in (p1, p2, p3): lst.sort(key=lambda r: r["age"])

# --- workbook ---
wb = openpyxl.Workbook(); wb.remove(wb.active)
BRAND = "2E4A62"
COLS = ["Заказчик", "Подрядчик (GC)", "Телефон GC", "Объект", "Назначение", "Тип работ",
        "Описание (RU)", "Площадь", "Бюджет работ (рынок) $", "Ваша прибыль 20% $", "Потолок затрат 80% $",
        "Дата", "Актуальность", "Веб-сверка", "Альт. телефон", "Что предложить"]
hf = Font(bold=True, color="FFFFFF"); thin = Side(style="thin", color="D0D0D0")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
green = Font(color="1E7A34", bold=True); amber = Font(color="9A6A00"); grey = Font(color="777777")
budf = Font(bold=True, color="10508A")
def build(title, data, headcolor, subtitle, rate_note):
    ws = wb.create_sheet(title)
    hint = f"{title} — {subtitle} · {len(data)} объектов · Miami-Dade, активные (CO нет), ≤180д · {rate_note} · звони по «Телефон GC»"
    ws.append([hint]); ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    ws["A1"].font = Font(bold=True, color="FFFFFF"); ws["A1"].fill = PatternFill("solid", fgColor=headcolor)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.append(COLS)
    for c in ws[2]:
        c.font = hf; c.fill = PatternFill("solid", fgColor=BRAND); c.border = bd
        c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    rowfill = PatternFill("solid", fgColor={"P1":"DCE9F7","P2":"FDE9D8","P3":"ECECEC"}[title[:2]])
    for r in data:
        ws.append([r["owner"], r["gc"], r["phone"], r["addr"], r["use"], r["work"], r["desc"],
                   r["sf"], r["bud"], r["prof"], r["cap"], r["date"], r["tier"], r["st"], r["found"], r["svc"]])
        row = ws[ws.max_row]
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True); c.border = bd; c.fill = rowfill
        row[8].font = budf; row[9].font = Font(bold=True, color="1E7A34")
        vc = row[13]
        if r["st"].startswith("✓"): vc.font = green
        elif r["st"].startswith("≠"): vc.font = amber
        elif r["st"].startswith("—"): vc.font = grey
    for i, wd in enumerate([18, 22, 14, 20, 13, 13, 24, 10, 16, 14, 15, 10, 11, 10, 13, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    ws.freeze_panes = "A3"; ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}{ws.max_row}"
    ws.row_dimensions[1].height = 26

build("P1 · Отделка", p1, "2E5E8C", "Покраска + гипсокартон + полы", "рыночная ставка $22/SF (полный пакет; прибыль 20% заложена)")
build("P2 · Уборка", p2, "B5651D", "Final clean после стройки", "рыночная ставка $0.55/SF, мин. $1500; прибыль 20% заложена")
build("P3 · Демонтаж", p3, "555555", "Демонтаж / вывоз мусора", "рыночная ставка $7/SF + вывоз ~$700/30yd; прибыль 20% заложена")

# --- вкладка РАСЦЕНКИ (рынок + источники) ---
rs = wb.create_sheet("Расценки (рынок)")
rs.column_dimensions["A"].width = 34; rs.column_dimensions["B"].width = 16
rs.column_dimensions["C"].width = 16; rs.column_dimensions["D"].width = 16; rs.column_dimensions["E"].width = 40
rs["A1"] = "РЫНОЧНЫЕ РАСЦЕНКИ Miami-Dade/Broward 2025-26 — субподряд (цена, которую платит GC)"
rs["A1"].font = Font(bold=True, size=13, color="FFFFFF"); rs["A1"].fill = PatternFill("solid", fgColor=BRAND)
rs.merge_cells("A1:E1")
hdr = ["Работа", "Низ $/SF", "Средн. $/SF", "Верх $/SF", "Источники / примечание"]
rs.append([]); rs.append(hdr)
for c in rs[3]:
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2E5E8C")
    c.alignment = Alignment(wrap_text=True, vertical="center")
rate_rows = [
    ("Покраска (коммерч. интерьер)", "1.50–2.00", "2.00–4.00", "4.00–6.00", "Javier's Painting, Solutions Painting FL, Homeyou Miami — Майами $3–6/SF (за floor SF)"),
    ("Гипсокартон (монтаж+шпат.+финиш)", "1.50–2.00", "2.00–3.00", "3.00–4.50", "Optima Construction, HomeGuide, Homewyse — за SF стены; ураган-код тянет верх"),
    ("Полы LVT/винил (монтаж+материал)", "6–9", "9–11", "11–14", "Terrapin CG, HomeGuide — коммерч. 20-mil glue-down по верху"),
    ("Полы плитка керамика/порцелан", "5–8", "8–12", "12–15+", "Aston Stuart Miami, HomeAdvisor — крупный формат/узор дороже"),
    ("Полы полированный бетон", "3–5", "5–8", "7–12", "HomeAdvisor, Terrapin — +$1.5–4.5/SF на подготовку основания"),
    ("ПАКЕТ отделки (гипс+краска+пол)", "12–16", "16–26", "26–40", "FABS Remodeling, Benchmark Building, Terrapin — за floor SF, малый build-out"),
    ("Финальная уборка после стройки", "0.10–0.30 (rough)", "0.30–0.75 (final)", "до 0.80 (Майами)", "Top Cleaning FL, Yorleny's, Angi — мин. работа ~$1500; окна $10–20/шт"),
    ("Демонтаж интерьера (не несущий)", "2–4", "4–8", "8–15", "Florida Demolition Experts, RemoveRight — вывоз 30yd ~$525–815; свалка +$2–3/SF"),
]
for r in rate_rows:
    rs.append(list(r))
    for c in rs[rs.max_row]: c.alignment = Alignment(wrap_text=True, vertical="top")
note = rs.max_row + 2
rs.cell(note, 1, "НАШИ ставки в файле (верхне-средние, прибыль 20% заложена):").font = Font(bold=True)
rs.cell(note+1, 1, "• P1 Отделка: $22/SF (полный пакет гипс+краска+пол). Если нужна 1 работа — цифру режь: только покраска ~$3–5/SF, только полы ~$9–12/SF, только гипс ~$2.5–4/SF.")
rs.cell(note+2, 1, "• P2 Уборка: $0.55/SF, минимум $1500 за объект.")
rs.cell(note+3, 1, "• P3 Демонтаж: $7/SF + вывоз ~$700 за 30-yd контейнер.")
rs.cell(note+4, 1, "• «Ваша прибыль 20%» = 20% от бюджета. «Потолок затрат 80%» = максимум на труд+материал, чтобы 20% реально остались.")
rs.cell(note+5, 1, "• ВАЖНО: краска/гипс часто считают за SF СТЕНЫ (≈2.5–3.5× от площади пола). Уточняй у GC метраж и что именно нужно — потом даёшь твёрдую цену.")
for rr in range(note, note+6):
    rs.cell(rr, 1).alignment = Alignment(wrap_text=True, vertical="top")
    rs.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5)

out = os.path.join(os.path.dirname(__file__), "..", "reports", "ISP-обзвон-ИТОГ-2026-08-26.xlsx")
wb.save(out)
print("wrote", out)
print(f"P1 отделка: {len(p1)} | P2 уборка: {len(p2)} | P3 демонтаж: {len(p3)}")
